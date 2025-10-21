"""
Pure openpyxl read/extract helpers for STAF V3.1.
IMPORTANT: We do NOT write/save the .xlsm with openpyxl (to avoid stripping shapes).
Writing comments is handled by xlwings COM in xlwings_comment.py.
"""
import re
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

def validate_ship_code(code: str) -> str:
    """
    Ensure ship code is exactly 2 alphabetic chars (e.g., 'GR').
    Raises ValueError if invalid.
    """
    code = (code or "").strip().upper()
    if len(code) != 2 or not code.isalpha():
        raise ValueError("Ship code must be exactly 2 alphabetic characters (e.g., 'GR').")
    return code

def load_workbooks_readonly(source_path: str, target_path: str):
    """
    Load Machine_Details (source) and STAF (target) workbooks in a read-safe way.
    - source: load normally
    - target: keep_vba=True and data_only=True so we read values without re-saving
    Note: Do not save with openpyxl to avoid shape loss.
    """
    try:
        source_wb = load_workbook(filename=source_path, data_only=True)
        target_wb = load_workbook(filename=target_path, keep_vba=True, data_only=True)
        return source_wb, target_wb
    except Exception as e:
        raise Exception(f"❌ Failed to load workbooks: {e}")

def build_comment_dict(sheet, ship_code: str) -> dict:
    """
    Build mapping: 'GR001' -> multiline comment text composed of key: value pairs.
    - Expects a header row in Row 1, with 'Position' among headers.
    - Rows start at 2.
    """
    headers = [str(c.value).strip() if c.value is not None else "" for c in sheet[1]]
    idx_position = None
    for i, h in enumerate(headers):
        if h.lower() == "position":
            idx_position = i
            break
    if idx_position is None:
        raise ValueError("❌ Source sheet must have a 'Position' header in row 1.")

    comment_dict = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        # Compose dictionary for this row
        row_data = {headers[i]: row[i] for i in range(len(headers))}
        pos_value = row_data.get("Position")
        if pos_value is None:
            continue
        try:
            pos_num = int(pos_value)
        except (TypeError, ValueError):
            continue
        pos_key = f"{ship_code}{pos_num:03d}"

        # Build multiline text
        lines = []
        for k, v in row_data.items():
            if v is None:
                continue
            lines.append(f"{k}: {v}")
        if not lines:
            continue
        comment_dict[pos_key] = "\\n".join(lines)
    return comment_dict

def extract_daily_metrics(target_wb, ship_code: str, machine_count: int, log_callback=None):
    """
    Find the 'DAILY COIN IN' and 'DAILY NET WIN' columns from the TOTALS sheet and
    return two dicts keyed by 'GR001', 'GR002', ...
    """
    try:
        sheet = target_wb["TOTALS"]
    except KeyError:
        raise ValueError("❌ 'TOTALS' sheet not found in target workbook.")

    coin_in_col = net_win_col = header_row = None

    for r in range(1, 20):
        for c in range(1, sheet.max_column + 1):
            value = sheet.cell(row=r, column=c).value
            if value:
                text = str(value).replace("\n", " ").strip().upper()
                text = re.sub(r"\s+", " ", text)
                if "DAILY COIN IN" in text:
                    coin_in_col = c
                    header_row = r
                elif "DAILY NET WIN" in text:
                    net_win_col = c
                    header_row = r
        if coin_in_col and net_win_col:
            break

    if not coin_in_col or not net_win_col:
        raise ValueError("❌ Couldn't find 'DAILY COIN IN' and 'DAILY NET WIN' headers.")

    coin_dict = {}
    netwin_dict = {}

    for i in range(1, machine_count + 1):
        row = header_row + i
        key = f"{ship_code}{i:03d}"
        coin_val = sheet.cell(row=row, column=coin_in_col).value or 0
        net_val = sheet.cell(row=row, column=net_win_col).value or 0
        coin_dict[key] = coin_val
        netwin_dict[key] = net_val

    if log_callback:
        log_callback(f"✅ Extracted Daily Coin-In and Net Win for {machine_count} positions.")

    return coin_dict, netwin_dict

def get_merged_range_bounds(sheet, row, col):
    """Return (min_row, min_col, max_row, max_col) if inside a merged range; else None."""
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return (min_row, min_col, max_row, max_col)
    return None

def get_value_merge_safe(sheet, row, col):
    """Return anchor value if cell is merged; else the cell's own value."""
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return sheet.cell(row=min_row, column=min_col).value
    return sheet.cell(row=row, column=col).value

def jump_over_merged(sheet, row, col, dr, dc):
    """Move in (dr, dc), skipping across current merged block."""
    bounds = get_merged_range_bounds(sheet, row, col)
    if bounds:
        min_row, min_col, max_row, max_col = bounds
        if dr < 0:
            row = min_row - 1
        elif dr > 0:
            row = max_row + 1
        if dc < 0:
            col = min_col - 1
        elif dc > 0:
            col = max_col + 1
        return row, col
    else:
        return row + dr, col + dc

def has_surrounding_position_number(sheet, row, col, expected_number: int) -> bool:
    """Check 8 neighbors (merge-safe) for an integer equal to expected_number."""
    bounds = get_merged_range_bounds(sheet, row, col)
    if bounds:
        min_row, min_col, max_row, max_col = bounds
        row = (min_row + max_row) // 2
        col = (min_col + max_col) // 2

    directions = [(-1,-1),(-1,0),(-1,1),(0,-1),(0,1),(1,-1),(1,0),(1,1)]
    for dr, dc in directions:
        r, c = jump_over_merged(sheet, row, col, dr, dc)
        if r < 1 or c < 1 or r > sheet.max_row or c > sheet.max_column:
            continue
        try:
            val = get_value_merge_safe(sheet, r, c)
            if isinstance(val, (int, float)) and int(val) == expected_number:
                return True
            if isinstance(val, str):
                cleaned = val.replace("$", "").replace(",", "").strip()
                if cleaned.isdigit() and int(cleaned) == expected_number:
                    return True
        except:
            continue
    return False

def detect_active_metric(floor_sheet, coin_dict, netwin_dict, log_callback=None) -> str:
    """Return 'coin_in' or 'net_win' after tallying hits in floor values."""
    coin_hits = 0
    net_hits = 0

    coin_values = set(round(float(v), 2) for v in coin_dict.values() if isinstance(v, (int, float)))
    net_values = set(round(float(v), 2) for v in netwin_dict.values() if isinstance(v, (int, float)))

    for row in floor_sheet.iter_rows(values_only=True):
        for val in row:
            try:
                num = round(float(val), 2)
                if num in coin_values:
                    coin_hits += 1
                elif num in net_values:
                    net_hits += 1
            except:
                continue

    if log_callback:
        log_callback(f"📊 Coin-In Matches: {coin_hits}, Net Win Matches: {net_hits}")

    if coin_hits > net_hits:
        return "coin_in"
    elif net_hits > coin_hits:
        return "net_win"
    else:
        raise ValueError("❌ Could not determine whether FLOOR PLAN is showing Coin-In or Net Win.")